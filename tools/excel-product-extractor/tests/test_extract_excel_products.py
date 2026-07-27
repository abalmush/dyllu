from __future__ import annotations

import contextlib
import io
import json
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import extract_excel_products as extractor


def png_bytes(
    color: tuple[int, ...] = (20, 80, 160),
    mode: str = "RGB",
    size: tuple[int, int] = (24, 24),
) -> bytes:
    output = io.BytesIO()
    Image.new(mode, size, color).save(output, format="PNG")
    return output.getvalue()


def source_jpeg_bytes(size: tuple[int, int] = (24, 24)) -> bytes:
    output = io.BytesIO()
    Image.new("RGB", size, (20, 80, 160)).save(output, format="JPEG", quality=92)
    return output.getvalue()


def add_image(
    worksheet,
    cell: str,
    data: bytes | None = None,
    size: tuple[int, int] = (24, 24),
) -> None:
    image = ExcelImage(io.BytesIO(data or png_bytes(size=size)))
    image.width, image.height = size
    worksheet.add_image(image, cell)


def add_product(
    worksheet,
    row: int,
    sku: str | None = "SKU-100",
    name: str = "Product",
    label: str = "Best seller",
    description: str = "Description",
    unit: str = "PCS",
    image_row: int | None = None,
    image_cell: str = "B",
    image_data: bytes | None = None,
) -> None:
    worksheet.cell(row=row, column=1, value=sku)
    worksheet.cell(row=row, column=3, value=label)
    worksheet.cell(row=row, column=4, value=name)
    worksheet.cell(row=row, column=5, value=description)
    worksheet.cell(row=row, column=6, value=unit)
    add_image(worksheet, f"{image_cell}{image_row or row}", image_data)


def save_workbook(path: Path, configure) -> None:
    workbook = Workbook()
    configure(workbook)
    workbook.save(path)


def invoke(arguments: list[str]) -> tuple[int, str]:
    stderr = io.StringIO()
    with contextlib.redirect_stderr(stderr):
        try:
            code = extractor.main(arguments)
        except SystemExit as exc:
            code = int(exc.code)
    return code, stderr.getvalue()


class ExtractExcelProductsTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.workbook_path = self.root / "catalogue.xlsx"
        self.output = self.root / "products"

    def tearDown(self) -> None:
        self.temporary_directory.cleanup()

    def run_catalogue(self, *extra: str) -> tuple[int, str]:
        return invoke(
            [str(self.workbook_path), "--output", str(self.output), *extra]
        )

    def manifest(self) -> list[dict[str, object]]:
        return json.loads((self.output / "products.json").read_text(encoding="utf-8"))

    def test_normal_extraction(self) -> None:
        save_workbook(self.workbook_path, lambda workbook: add_product(workbook.active, 2))
        code, error = self.run_catalogue()
        self.assertEqual((code, error), (extractor.EXIT_OK, ""))
        self.assertTrue((self.output / "SKU-100.jpeg").is_file())
        payload = json.loads((self.output / "SKU-100.json").read_text(encoding="utf-8"))
        self.assertEqual(payload["source"]["row"], 2)

    def test_multiple_products(self) -> None:
        def configure(workbook: Workbook) -> None:
            add_product(workbook.active, 2, sku="SKU-100")
            add_product(workbook.active, 3, sku="SKU-200")

        save_workbook(self.workbook_path, configure)
        code, _ = self.run_catalogue()
        self.assertEqual(code, extractor.EXIT_OK)
        self.assertEqual([item["sku"] for item in self.manifest()], ["SKU-100", "SKU-200"])

    def test_multiline_description_is_preserved(self) -> None:
        description = "First line\nVoltage: 20V\nLast line"
        save_workbook(
            self.workbook_path,
            lambda workbook: add_product(
                workbook.active, 2, description=description
            ),
        )
        self.run_catalogue()
        self.assertEqual(self.manifest()[0]["description"], description)

    def test_merged_sku_cell_is_resolved(self) -> None:
        def configure(workbook: Workbook) -> None:
            sheet = workbook.active
            sheet.merge_cells("A2:A3")
            sheet["A2"] = "MERGED-1"
            add_product(sheet, 3, sku=None)

        save_workbook(self.workbook_path, configure)
        code, _ = self.run_catalogue()
        self.assertEqual(code, extractor.EXIT_OK)
        self.assertEqual(self.manifest()[0]["sku"], "MERGED-1")

    def test_transparency_is_flattened_to_white(self) -> None:
        transparent = png_bytes((255, 0, 0, 0), mode="RGBA")
        save_workbook(
            self.workbook_path,
            lambda workbook: add_product(
                workbook.active, 2, image_data=transparent
            ),
        )
        self.run_catalogue("--jpeg-quality", "100")
        with Image.open(self.output / "SKU-100.jpeg") as extracted:
            red, green, blue = extracted.convert("RGB").getpixel((10, 10))
        self.assertGreater(min(red, green, blue), 245)

    def test_upscale_doubles_image_dimensions(self) -> None:
        save_workbook(self.workbook_path, lambda workbook: add_product(workbook.active, 2))
        code, _ = self.run_catalogue("--upscale", "2")
        self.assertEqual(code, extractor.EXIT_OK)
        with Image.open(self.output / "SKU-100.jpeg") as extracted:
            self.assertEqual(extracted.size, (48, 48))
            self.assertEqual(extracted.mode, "RGB")

    def test_original_jpeg_is_preserved_without_upscaling(self) -> None:
        source = source_jpeg_bytes()
        save_workbook(
            self.workbook_path,
            lambda workbook: add_product(workbook.active, 2, image_data=source),
        )
        code, _ = self.run_catalogue()
        self.assertEqual(code, extractor.EXIT_OK)
        self.assertEqual((self.output / "SKU-100.jpeg").read_bytes(), source)

    def test_invalid_upscale_factor(self) -> None:
        code, error = self.run_catalogue("--upscale", "0")
        self.assertEqual(code, extractor.EXIT_USAGE)
        self.assertIn("upscale factor", error)

    def test_missing_sku_is_warned_and_returns_no_products(self) -> None:
        save_workbook(
            self.workbook_path,
            lambda workbook: add_product(workbook.active, 2, sku=None),
        )
        code, _ = self.run_catalogue()
        self.assertEqual(code, extractor.EXIT_NO_PRODUCTS)
        self.assertIn("no nearby SKU", (self.output / "warnings.txt").read_text())

    def test_image_one_row_away_matches_metadata(self) -> None:
        save_workbook(
            self.workbook_path,
            lambda workbook: add_product(workbook.active, 3, image_row=2),
        )
        code, _ = self.run_catalogue("--search-radius", "1")
        self.assertEqual(code, extractor.EXIT_OK)
        self.assertEqual(self.manifest()[0]["source"]["imageAnchorRow"], 2)
        self.assertEqual(self.manifest()[0]["source"]["row"], 3)

    def test_duplicate_skus_receive_suffixes(self) -> None:
        def configure(workbook: Workbook) -> None:
            add_product(workbook.active, 2, sku="DUP-1")
            add_product(workbook.active, 3, sku="DUP-1")

        save_workbook(self.workbook_path, configure)
        self.run_catalogue()
        self.assertTrue((self.output / "DUP-1.jpeg").is_file())
        self.assertTrue((self.output / "DUP-1-2.jpeg").is_file())

    def test_invalid_workbook_path(self) -> None:
        code, error = self.run_catalogue()
        self.assertEqual(code, extractor.EXIT_WORKBOOK)
        self.assertIn("does not exist", error)

    def test_malformed_workbook(self) -> None:
        self.workbook_path.write_bytes(b"not an Excel workbook")
        code, error = self.run_catalogue()
        self.assertEqual(code, extractor.EXIT_WORKBOOK)
        self.assertIn("could not read workbook", error)

    def test_unsupported_xls(self) -> None:
        legacy = self.root / "catalogue.xls"
        legacy.write_bytes(b"legacy")
        code, error = invoke([str(legacy), "--output", str(self.output)])
        self.assertEqual(code, extractor.EXIT_WORKBOOK)
        self.assertIn("convert the file to .xlsx", error)

    def test_unknown_sheet(self) -> None:
        save_workbook(self.workbook_path, lambda workbook: add_product(workbook.active, 2))
        code, error = self.run_catalogue("--sheet", "Missing")
        self.assertEqual(code, extractor.EXIT_WORKBOOK)
        self.assertIn("unknown sheet", error)

    def test_invalid_column_argument(self) -> None:
        code, error = self.run_catalogue("--sku-col", "A1")
        self.assertEqual(code, extractor.EXIT_USAGE)
        self.assertIn("invalid Excel column", error)

    def test_duplicate_column_arguments_are_invalid(self) -> None:
        code, error = self.run_catalogue("--sku-col", "C")
        self.assertEqual(code, extractor.EXIT_USAGE)
        self.assertIn("distinct columns", error)

    def test_workbook_with_no_images(self) -> None:
        def configure(workbook: Workbook) -> None:
            workbook.active["A2"] = "SKU-100"

        save_workbook(self.workbook_path, configure)
        code, _ = self.run_catalogue()
        self.assertEqual(code, extractor.EXIT_NO_PRODUCTS)
        self.assertIn("No embedded images", (self.output / "warnings.txt").read_text())

    def test_images_without_matching_products(self) -> None:
        def configure(workbook: Workbook) -> None:
            add_image(workbook.active, "B10")

        save_workbook(self.workbook_path, configure)
        code, _ = self.run_catalogue("--search-radius", "1")
        self.assertEqual(code, extractor.EXIT_NO_PRODUCTS)
        self.assertEqual(self.manifest(), [])

    def test_zip_generation(self) -> None:
        save_workbook(self.workbook_path, lambda workbook: add_product(workbook.active, 2))
        code, _ = self.run_catalogue("--zip")
        self.assertEqual(code, extractor.EXIT_OK)
        with zipfile.ZipFile(Path(f"{self.output}.zip")) as archive:
            self.assertEqual(
                set(archive.namelist()),
                {"SKU-100.jpeg", "SKU-100.json", "products.json", "warnings.txt"},
            )

    def test_non_empty_output_directory_fails_closed(self) -> None:
        save_workbook(self.workbook_path, lambda workbook: add_product(workbook.active, 2))
        self.output.mkdir()
        (self.output / "keep.txt").write_text("keep", encoding="utf-8")
        code, error = self.run_catalogue()
        self.assertEqual(code, extractor.EXIT_OUTPUT)
        self.assertIn("must be empty", error)
        self.assertEqual((self.output / "keep.txt").read_text(encoding="utf-8"), "keep")

    def test_invalid_filename_characters_are_sanitized(self) -> None:
        save_workbook(
            self.workbook_path,
            lambda workbook: add_product(workbook.active, 2, sku='SKU/12:*?"'),
        )
        self.run_catalogue()
        image_name = self.manifest()[0]["imageFile"]
        self.assertFalse(set('<>:"/\\|?*') & set(image_name))
        self.assertTrue((self.output / image_name).is_file())

    def test_non_ascii_description(self) -> None:
        description = "Șurubelniță fără fir\n中文说明"
        save_workbook(
            self.workbook_path,
            lambda workbook: add_product(
                workbook.active, 2, description=description
            ),
        )
        self.run_catalogue()
        self.assertEqual(self.manifest()[0]["description"], description)

    def test_multiple_sheets(self) -> None:
        def configure(workbook: Workbook) -> None:
            first = workbook.active
            first.title = "First"
            add_product(first, 2, sku="FIRST-1")
            second = workbook.create_sheet("Second")
            add_product(second, 2, sku="SECOND-1")

        save_workbook(self.workbook_path, configure)
        code, _ = self.run_catalogue("--sheet", "First", "--sheet", "Second")
        self.assertEqual(code, extractor.EXIT_OK)
        self.assertEqual({item["sku"] for item in self.manifest()}, {"FIRST-1", "SECOND-1"})

    def test_row_offset(self) -> None:
        save_workbook(
            self.workbook_path,
            lambda workbook: add_product(workbook.active, 3, image_row=2),
        )
        code, _ = self.run_catalogue("--row-offset", "1", "--search-radius", "0")
        self.assertEqual(code, extractor.EXIT_OK)

    def test_final_fallback_scans_nearby_cells(self) -> None:
        def configure(workbook: Workbook) -> None:
            sheet = workbook.active
            sheet["D2"] = "ALT-200"
            add_image(sheet, "B2")

        save_workbook(self.workbook_path, configure)
        code, _ = self.run_catalogue("--name-col", "G")
        self.assertEqual(code, extractor.EXIT_OK)
        self.assertEqual(self.manifest()[0]["sku"], "ALT-200")

    def test_largest_image_is_selected_for_a_product_row(self) -> None:
        def configure(workbook: Workbook) -> None:
            sheet = workbook.active
            add_product(sheet, 2)
            add_image(sheet, "A2", size=(6, 6))

        save_workbook(self.workbook_path, configure)
        code, _ = self.run_catalogue()
        self.assertEqual(code, extractor.EXIT_OK)
        self.assertEqual(len(self.manifest()), 1)
        self.assertIn("a primary image", (self.output / "warnings.txt").read_text())

    def test_exact_non_zero_exit_codes(self) -> None:
        missing_code, _ = self.run_catalogue()
        self.workbook_path.write_bytes(b"broken")
        malformed_code, _ = self.run_catalogue()
        invalid_code, _ = self.run_catalogue("--search-radius", "-1")
        self.assertEqual(
            (missing_code, malformed_code, invalid_code),
            (extractor.EXIT_WORKBOOK, extractor.EXIT_WORKBOOK, extractor.EXIT_USAGE),
        )


if __name__ == "__main__":
    unittest.main()
