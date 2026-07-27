#!/usr/bin/env python3

from __future__ import annotations

import argparse
import io
import json
import re
import sys
import unicodedata
import warnings as python_warnings
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from PIL import Image, ImageFilter, ImageOps, UnidentifiedImageError


EXIT_OK = 0
EXIT_USAGE = 2
EXIT_WORKBOOK = 3
EXIT_NO_PRODUCTS = 4
EXIT_OUTPUT = 5


class ConfigurationError(Exception):
    pass


class WorkbookProcessingError(Exception):
    pass


class OutputError(Exception):
    pass


@dataclass(frozen=True)
class ColumnMapping:
    sku: int
    label: int
    name: int
    description: int
    unit: int


@dataclass(frozen=True)
class ExtractionConfig:
    workbook_path: Path
    output_dir: Path
    sheets: tuple[str, ...]
    columns: ColumnMapping
    row_offset: int
    search_radius: int
    jpeg_quality: int
    upscale: int
    make_zip: bool


@dataclass(frozen=True)
class ImageCandidate:
    sheet_name: str
    anchor_row: int
    anchor_col: int
    product_row: int
    sku: str
    data: bytes
    selection_area: int
    image_number: int


@dataclass(frozen=True)
class Product:
    sku: str
    name: str
    label: str
    description: str
    unit: str
    workbook_name: str
    sheet_name: str
    row: int
    image_anchor_row: int
    image_data: bytes


def column_argument(value: str) -> str:
    normalized = value.strip().upper()
    try:
        index = column_index_from_string(normalized)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"invalid Excel column: {value!r}") from exc
    if not 1 <= index <= 16384:
        raise argparse.ArgumentTypeError(f"Excel column is out of range: {value!r}")
    return get_column_letter(index)


def non_negative_integer(value: str) -> int:
    try:
        parsed = int(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"expected an integer, got {value!r}") from exc
    if parsed < 0:
        raise argparse.ArgumentTypeError("value must be zero or greater")
    return parsed


def jpeg_quality_argument(value: str) -> int:
    try:
        parsed = int(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"expected an integer, got {value!r}") from exc
    if not 1 <= parsed <= 100:
        raise argparse.ArgumentTypeError("JPEG quality must be between 1 and 100")
    return parsed


def upscale_argument(value: str) -> int:
    try:
        parsed = int(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"expected an integer, got {value!r}") from exc
    if not 1 <= parsed <= 4:
        raise argparse.ArgumentTypeError("upscale factor must be between 1 and 4")
    return parsed


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract product images and metadata from an Excel catalogue."
    )
    parser.add_argument("workbook", type=Path, help="Path to an .xlsx or .xlsm workbook")
    parser.add_argument("--output", type=Path, default=Path("products"))
    parser.add_argument("--sheet", action="append", default=[], help="Sheet to process; repeatable")
    parser.add_argument("--sku-col", type=column_argument, default="A")
    parser.add_argument("--label-col", type=column_argument, default="C")
    parser.add_argument("--name-col", type=column_argument, default="D")
    parser.add_argument("--description-col", type=column_argument, default="E")
    parser.add_argument("--unit-col", type=column_argument, default="F")
    parser.add_argument("--row-offset", type=int, default=0)
    parser.add_argument("--search-radius", type=non_negative_integer, default=2)
    parser.add_argument("--jpeg-quality", type=jpeg_quality_argument, default=95)
    parser.add_argument("--upscale", type=upscale_argument, default=1)
    parser.add_argument("--zip", action="store_true", dest="make_zip")
    return parser


def config_from_args(args: argparse.Namespace) -> ExtractionConfig:
    names = [args.sku_col, args.label_col, args.name_col, args.description_col, args.unit_col]
    if len(set(names)) != len(names):
        raise ConfigurationError("column arguments must refer to distinct columns")
    if any(not sheet.strip() for sheet in args.sheet):
        raise ConfigurationError("sheet names cannot be empty")
    sheets = tuple(dict.fromkeys(args.sheet))
    return ExtractionConfig(
        workbook_path=args.workbook.expanduser(),
        output_dir=args.output.expanduser(),
        sheets=sheets,
        columns=ColumnMapping(
            sku=column_index_from_string(args.sku_col),
            label=column_index_from_string(args.label_col),
            name=column_index_from_string(args.name_col),
            description=column_index_from_string(args.description_col),
            unit=column_index_from_string(args.unit_col),
        ),
        row_offset=args.row_offset,
        search_radius=args.search_radius,
        jpeg_quality=args.jpeg_quality,
        upscale=args.upscale,
        make_zip=args.make_zip,
    )


def validate_workbook_path(path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        raise WorkbookProcessingError(
            "legacy .xls files are not supported by openpyxl; convert the file to .xlsx first"
        )
    if suffix not in {".xlsx", ".xlsm"}:
        raise ConfigurationError("workbook must use the .xlsx or .xlsm extension")
    if not path.is_file():
        raise WorkbookProcessingError(f"workbook does not exist or is not a file: {path}")


def validate_output_target(config: ExtractionConfig) -> None:
    try:
        if config.output_dir.exists():
            if not config.output_dir.is_dir():
                raise OutputError(
                    f"output path exists and is not a directory: {config.output_dir}"
                )
            if any(config.output_dir.iterdir()):
                raise OutputError(
                    f"output directory must be empty: {config.output_dir}"
                )
        if config.make_zip:
            zip_path = Path(f"{config.output_dir}.zip")
            if zip_path.exists():
                raise OutputError(f"ZIP output already exists: {zip_path}")
    except OSError as exc:
        raise OutputError(f"could not inspect output path: {exc}") from exc


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    return text.replace("\r\n", "\n").replace("\r", "\n").strip()


def merged_cell_value(worksheet, row: int, column: int) -> object:
    if row < 1 or column < 1:
        return None
    value = worksheet.cell(row=row, column=column).value
    if value is not None:
        return value
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return worksheet.cell(
                row=merged_range.min_row, column=merged_range.min_col
            ).value
    return None


def looks_like_sku(value: object) -> bool:
    text = normalize_text(value)
    if not 3 <= len(text) <= 80 or any(character.isspace() for character in text):
        return False
    if "@" in text or not any(character.isalpha() for character in text):
        return False
    if not any(character.isdigit() for character in text):
        return False
    allowed_punctuation = set("._:/\\*?+#-")
    return all(character.isalnum() or character in allowed_punctuation for character in text)


def nearby_rows(target_row: int, radius: int, max_row: int) -> list[int]:
    rows: list[int] = []
    if 1 <= target_row <= max_row:
        rows.append(target_row)
    for distance in range(1, radius + 1):
        for row in (target_row - distance, target_row + distance):
            if 1 <= row <= max_row:
                rows.append(row)
    return rows


def match_product_row(
    worksheet,
    anchor_row: int,
    anchor_col: int,
    sku_column: int,
    row_offset: int,
    search_radius: int,
) -> tuple[int, str] | None:
    target_row = anchor_row + row_offset
    if 1 <= target_row <= worksheet.max_row:
        direct_sku = normalize_text(merged_cell_value(worksheet, target_row, sku_column))
        if direct_sku:
            return target_row, direct_sku

    rows = nearby_rows(target_row, search_radius, worksheet.max_row)
    for row in rows:
        if row == target_row:
            continue
        candidate = merged_cell_value(worksheet, row, sku_column)
        if looks_like_sku(candidate):
            return row, normalize_text(candidate)

    minimum_column = max(1, min(sku_column, anchor_col) - 3)
    maximum_column = min(worksheet.max_column, max(sku_column, anchor_col) + 3)
    for row in rows:
        for column in range(minimum_column, maximum_column + 1):
            if column == sku_column:
                continue
            candidate = merged_cell_value(worksheet, row, column)
            if looks_like_sku(candidate):
                return row, normalize_text(candidate)
    return None


def image_anchor(image) -> tuple[int, int]:
    marker = getattr(image.anchor, "_from", None)
    if marker is None:
        raise ValueError("unsupported image anchor")
    return marker.row + 1, marker.col + 1


def validate_image(data: bytes) -> int:
    with Image.open(io.BytesIO(data)) as source:
        source.load()
        width, height = source.size
    if width < 1 or height < 1:
        raise ValueError("image has invalid dimensions")
    return width * height


def image_selection_area(image, pixel_area: int) -> int:
    anchor = image.anchor
    extent = getattr(anchor, "ext", None)
    if extent is not None:
        width = getattr(extent, "cx", 0) / 9525
        height = getattr(extent, "cy", 0) / 9525
        if width > 0 and height > 0:
            return round(width * height)
    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)
    if start is not None and end is not None:
        width = (end.col - start.col) * 64 + (end.colOff - start.colOff) / 9525
        height = (end.row - start.row) * 20 + (end.rowOff - start.rowOff) / 9525
        if width > 0 and height > 0:
            return round(width * height)
    return pixel_area


def extract_candidates(
    worksheets,
    config: ExtractionConfig,
    warnings: list[str],
) -> list[ImageCandidate]:
    candidates: list[ImageCandidate] = []
    image_count = 0
    for worksheet in worksheets:
        images = list(getattr(worksheet, "_images", []))
        for image_number, image in enumerate(images, start=1):
            image_count += 1
            try:
                anchor_row, anchor_col = image_anchor(image)
            except (AttributeError, TypeError, ValueError) as exc:
                warnings.append(
                    f"{worksheet.title}: image {image_number} skipped: {exc}"
                )
                continue
            match = match_product_row(
                worksheet,
                anchor_row,
                anchor_col,
                config.columns.sku,
                config.row_offset,
                config.search_radius,
            )
            if match is None:
                warnings.append(
                    f"{worksheet.title}: image {image_number} at row {anchor_row} skipped: "
                    "no nearby SKU could be matched"
                )
                continue
            product_row, sku = match
            try:
                data = image._data()
                pixel_area = validate_image(data)
                selection_area = image_selection_area(image, pixel_area)
            except Exception as exc:
                warnings.append(
                    f"{worksheet.title}: image {image_number} at row {anchor_row} skipped: "
                    f"malformed or unreadable image ({exc})"
                )
                continue
            candidates.append(
                ImageCandidate(
                    sheet_name=worksheet.title,
                    anchor_row=anchor_row,
                    anchor_col=anchor_col,
                    product_row=product_row,
                    sku=sku,
                    data=data,
                    selection_area=selection_area,
                    image_number=image_number,
                )
            )
    if image_count == 0:
        warnings.append("No embedded images were found in the selected sheets")
    return candidates


def select_primary_images(
    candidates: Sequence[ImageCandidate], warnings: list[str]
) -> list[ImageCandidate]:
    grouped: dict[tuple[str, int], list[ImageCandidate]] = defaultdict(list)
    group_order: list[tuple[str, int]] = []
    for candidate in candidates:
        key = (candidate.sheet_name, candidate.product_row)
        if key not in grouped:
            group_order.append(key)
        grouped[key].append(candidate)

    selected: list[ImageCandidate] = []
    for key in group_order:
        group = grouped[key]
        primary = max(group, key=lambda item: (item.selection_area, -item.image_number))
        selected.append(primary)
        for candidate in group:
            if candidate is primary:
                continue
            warnings.append(
                f"{candidate.sheet_name}: image {candidate.image_number} at row "
                f"{candidate.anchor_row} skipped: a primary image is already "
                f"matched to product row {candidate.product_row}"
            )
    return selected


def build_products(
    workbook_name: str,
    workbook,
    candidates: Sequence[ImageCandidate],
    columns: ColumnMapping,
) -> list[Product]:
    products: list[Product] = []
    for candidate in candidates:
        worksheet = workbook[candidate.sheet_name]
        row = candidate.product_row
        products.append(
            Product(
                sku=candidate.sku,
                name=normalize_text(merged_cell_value(worksheet, row, columns.name)),
                label=normalize_text(merged_cell_value(worksheet, row, columns.label)),
                description=normalize_text(
                    merged_cell_value(worksheet, row, columns.description)
                ),
                unit=normalize_text(merged_cell_value(worksheet, row, columns.unit)),
                workbook_name=workbook_name,
                sheet_name=candidate.sheet_name,
                row=row,
                image_anchor_row=candidate.anchor_row,
                image_data=candidate.data,
            )
        )
    return products


def sanitize_sku(sku: str) -> str:
    normalized = unicodedata.normalize("NFKC", sku).strip()
    normalized = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", "_", normalized)
    normalized = re.sub(r"\s+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip(" ._")
    if not normalized:
        normalized = "product"
    reserved = {"CON", "PRN", "AUX", "NUL"}
    reserved.update(f"COM{number}" for number in range(1, 10))
    reserved.update(f"LPT{number}" for number in range(1, 10))
    if normalized.upper() in reserved:
        normalized = f"_{normalized}"
    if normalized.casefold() == "products":
        normalized = "_products"
    return normalized[:180].rstrip(" .") or "product"


def jpeg_bytes(data: bytes, quality: int, upscale: int) -> bytes:
    with Image.open(io.BytesIO(data)) as source:
        source.load()
        orientation = source.getexif().get(274, 1)
        if (
            upscale == 1
            and source.format == "JPEG"
            and source.mode == "RGB"
            and orientation == 1
        ):
            return data
        source = ImageOps.exif_transpose(source)
        has_alpha = source.mode in {"RGBA", "LA"} or (
            source.mode == "P" and "transparency" in source.info
        )
        if has_alpha:
            rgba = source.convert("RGBA")
            rgb = Image.new("RGB", rgba.size, "white")
            rgb.paste(rgba, mask=rgba.getchannel("A"))
        else:
            rgb = source.convert("RGB")
        if upscale > 1:
            size = (round(rgb.width * upscale), round(rgb.height * upscale))
            rgb = rgb.resize(size, Image.Resampling.LANCZOS)
            rgb = rgb.filter(
                ImageFilter.UnsharpMask(radius=1.2, percent=100, threshold=3)
            )
        output = io.BytesIO()
        rgb.save(
            output,
            format="JPEG",
            quality=quality,
            optimize=True,
            progressive=True,
            subsampling=0,
        )
        return output.getvalue()


def product_payload(product: Product) -> dict[str, object]:
    return {
        "sku": product.sku,
        "name": product.name,
        "label": product.label,
        "description": product.description,
        "unit": product.unit,
        "source": {
            "workbook": product.workbook_name,
            "sheet": product.sheet_name,
            "row": product.row,
            "imageAnchorRow": product.image_anchor_row,
        },
    }


def write_json(path: Path, payload: object) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def write_outputs(
    products: Sequence[Product], config: ExtractionConfig, warnings: list[str]
) -> None:
    try:
        config.output_dir.mkdir(parents=True, exist_ok=True)
        manifest: list[dict[str, object]] = []
        filename_counts: dict[str, int] = defaultdict(int)
        for product in products:
            base_name = sanitize_sku(product.sku)
            collision_key = base_name.casefold()
            filename_counts[collision_key] += 1
            occurrence = filename_counts[collision_key]
            unique_name = base_name if occurrence == 1 else f"{base_name}-{occurrence}"
            image_filename = f"{unique_name}.jpeg"
            json_filename = f"{unique_name}.json"
            image_data = jpeg_bytes(
                product.image_data, config.jpeg_quality, config.upscale
            )
            (config.output_dir / image_filename).write_bytes(image_data)
            payload = product_payload(product)
            write_json(config.output_dir / json_filename, payload)
            manifest.append(
                {
                    **payload,
                    "imageFile": image_filename,
                    "jsonFile": json_filename,
                }
            )
        write_json(config.output_dir / "products.json", manifest)
        warning_text = "\n".join(warnings)
        if warning_text:
            warning_text += "\n"
        (config.output_dir / "warnings.txt").write_text(warning_text, encoding="utf-8")
        if config.make_zip:
            zip_path = Path(f"{config.output_dir}.zip")
            with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                for path in sorted(config.output_dir.iterdir()):
                    if path.is_file():
                        archive.write(path, arcname=path.name)
    except (OSError, UnidentifiedImageError, ValueError, zipfile.BadZipFile) as exc:
        raise OutputError(f"could not write extraction output: {exc}") from exc


def process_workbook(config: ExtractionConfig) -> int:
    validate_workbook_path(config.workbook_path)
    validate_output_target(config)
    load_messages: list[str] = []
    try:
        with python_warnings.catch_warnings(record=True) as caught:
            python_warnings.simplefilter("always")
            workbook = load_workbook(
                config.workbook_path,
                data_only=True,
                read_only=False,
                keep_vba=config.workbook_path.suffix.lower() == ".xlsm",
            )
        load_messages.extend(f"Workbook warning: {warning.message}" for warning in caught)
    except (InvalidFileException, OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        raise WorkbookProcessingError(f"could not read workbook: {exc}") from exc
    except Exception as exc:
        raise WorkbookProcessingError(
            f"could not read workbook ({type(exc).__name__}): {exc}"
        ) from exc

    try:
        if config.sheets:
            missing = [name for name in config.sheets if name not in workbook.sheetnames]
            if missing:
                raise WorkbookProcessingError(
                    "unknown sheet(s): "
                    + ", ".join(missing)
                    + "; available sheets: "
                    + ", ".join(workbook.sheetnames)
                )
            worksheets = [workbook[name] for name in config.sheets]
        else:
            worksheets = list(workbook.worksheets)

        candidates = extract_candidates(worksheets, config, load_messages)
        selected = select_primary_images(candidates, load_messages)
        products = build_products(
            config.workbook_path.name, workbook, selected, config.columns
        )
        if not products:
            load_messages.append("No products were extracted")
        write_outputs(products, config, load_messages)
        return EXIT_OK if products else EXIT_NO_PRODUCTS
    finally:
        workbook.close()


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        config = config_from_args(args)
        return process_workbook(config)
    except ConfigurationError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return EXIT_USAGE
    except WorkbookProcessingError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return EXIT_WORKBOOK
    except OutputError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return EXIT_OUTPUT


if __name__ == "__main__":
    raise SystemExit(main())
